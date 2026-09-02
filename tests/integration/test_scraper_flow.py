"""Testy integracyjne – przepływ z mockowanym HTTP i zapisem Excel."""

from __future__ import annotations

from datetime import date
from unittest.mock import MagicMock, patch

import pytest
from openpyxl import load_workbook

import neueroeffnung_scraper as scraper
from tests.conftest import MockHttpResponse, build_mock_session

pytestmark = pytest.mark.integration


class TestCollectCategoryRecords:
    def test_collects_market_records_from_mocked_pages(
        self, load_fixture, silent_logger, monkeypatch
    ):
        monkeypatch.setattr(scraper, "REQUEST_DELAY_SEC", 0)
        monkeypatch.setattr(scraper, "MAX_PAGES_PER_CATEGORY", 2)

        route_map = [
            ("?page=2", load_fixture("list_page_plain_cards.html")),
            ("/detmold/lidl", load_fixture("detail_rewe_esch.html")),
            ("/muenchen/edeka-van-dungen", load_fixture("detail_edeka.html")),
            ("/moenchengladbach/rewe-esch", load_fixture("detail_rewe_esch.html")),
            ("/frankfurt-am-main/rewe-test", load_fixture("detail_rewe_esch.html")),
            ("/branche/supermaerkte", load_fixture("list_page_linked.html")),
        ]
        search_map = {
            "lidl": load_fixture("search_result_lidl.html"),
            "edeka": load_fixture("search_result_edeka.html"),
        }
        session = build_mock_session(route_map, search_map)

        records = scraper.collect_category_records(
            session=session,
            base_url=f"{scraper.BASE_URL}/branche/supermaerkte",
            cache={},
            logger=silent_logger,
        )

        assert len(records) == 4
        assert all(scraper.is_opening_date_in_range(r.data_otwarcia) for r in records)
        assert records[0].nazwa_firmy == "REWE Esch"
        assert records[0].typ_wpisu == "Reopening"
        assert records[0].adres == "Beckrather Straße 39, 41189 Mönchengladbach"
        assert "Kernsanierung" in records[0].informacja
        assert records[2].nazwa_firmy == "EDEKA van Dungen"
        assert records[2].adres == "Schellingstraße 11, 80331 München"
        assert scraper.is_incomplete_address(records[2].adres) is False

    def test_skips_records_outside_opening_date_range(
        self, load_fixture, silent_logger, monkeypatch
    ):
        monkeypatch.setattr(scraper, "REQUEST_DELAY_SEC", 0)
        monkeypatch.setattr(scraper, "MAX_PAGES_PER_CATEGORY", 1)

        route_map = [
            ("/berlin/zukunft-markt", load_fixture("detail_zukunft_markt.html")),
            ("/branche/supermaerkte", load_fixture("list_page_out_of_range.html")),
        ]
        session = build_mock_session(route_map)

        records = scraper.collect_category_records(
            session=session,
            base_url=f"{scraper.BASE_URL}/branche/supermaerkte",
            cache={},
            logger=silent_logger,
        )

        assert records == []


class TestWriteExcelIntegration:
    def test_writes_three_sheets_with_expected_columns(
        self, silent_logger, tmp_path
    ):
        output_path = tmp_path / "wynik.xlsx"
        long_adres = "A" * scraper.MAX_ADDRESS_LENGTH
        sheets = {
            "Markety": [
                scraper.Record("REWE Esch", "Beckrather Straße 39, 41189 Mönchengladbach", "", "03.09.2026"),
            ],
            "Restauracje": [
                scraper.Record("McDonald's", long_adres, "01.08.2026", "15.10.2026"),
            ],
            "Drogerie": [],
            "Centra handlowe": [],
        }

        scraper.write_excel(sheets, [], output_path, silent_logger)

        wb = load_workbook(output_path)
        assert "Harmonogram" in wb.sheetnames
        assert "Według regionu" in wb.sheetnames
        assert "Raport braków" in wb.sheetnames
        assert "Pominięte" in wb.sheetnames
        assert wb["Restauracje"]["B2"].value == long_adres
        assert len(wb["Restauracje"]["B2"].value) == 650


class TestFetchDetailUsesCache:
    def test_does_not_call_http_when_cached(self, silent_logger):
        session = MagicMock()
        cache = {
            "https://example.com/entry": {
                "adres": "Teststraße 1, 12345 Berlin",
                "data_otwarcia": "03.09.2026",
                "data_zamkniecia": "01.06.2026",
                "informacja": "Opis testowego sklepu.",
            }
        }

        adres, otwarcie, zamkniecie, informacja = scraper.fetch_detail(
            session, "https://example.com/entry", cache, silent_logger
        )

        session.get.assert_not_called()
        assert adres == "Teststraße 1, 12345 Berlin"
        assert otwarcie == "03.09.2026"
        assert informacja == "Opis testowego sklepu."


class TestPrepareFreshOutput:
    def test_removes_previous_result_files(self, tmp_path, silent_logger, monkeypatch):
        monkeypatch.setattr(scraper, "OUTPUT_FILE", tmp_path / "out.xlsx")
        monkeypatch.setattr(scraper, "JSON_OUTPUT_FILE", tmp_path / "out.json")
        monkeypatch.setattr(scraper, "VALIDATION_REPORT_FILE", tmp_path / "raport.json")

        (tmp_path / "out.xlsx").write_text("stary excel", encoding="utf-8")
        (tmp_path / "out.json").write_text("{}", encoding="utf-8")
        (tmp_path / "raport.json").write_text("{}", encoding="utf-8")

        scraper.prepare_fresh_output_files(silent_logger)

        assert not (tmp_path / "out.xlsx").exists()
        assert not (tmp_path / "out.json").exists()
        assert not (tmp_path / "raport.json").exists()


class TestRunScraperEndToEnd:
    def test_run_scraper_produces_excel(self, load_fixture, tmp_path, monkeypatch):
        monkeypatch.setattr(scraper, "REQUEST_DELAY_SEC", 0)
        monkeypatch.setattr(scraper, "MAX_PAGES_PER_CATEGORY", 1)
        monkeypatch.setattr(scraper, "OUTPUT_FILE", tmp_path / "out.xlsx")
        monkeypatch.setattr(scraper, "JSON_OUTPUT_FILE", tmp_path / "out.json")
        monkeypatch.setattr(scraper, "VALIDATION_REPORT_FILE", tmp_path / "raport.json")
        monkeypatch.setattr(scraper, "CACHE_FILE", tmp_path / "cache.json")
        monkeypatch.setattr(scraper, "LOG_FILE", tmp_path / "test.log")

        market_html = load_fixture("list_page_linked.html")
        detail_html = load_fixture("detail_rewe_esch.html")
        gastro_html = load_fixture("list_page_plain_cards.html")

        def fake_get(url, **kwargs):
            if "supermaerkte" in url and "page=" not in url:
                return MockHttpResponse(market_html)
            if "gastronomie" in url:
                return MockHttpResponse(gastro_html)
            if "/suche/" in url:
                return MockHttpResponse(load_fixture("search_result_lidl.html"))
            if "rewe-esch" in url or "rewe-test" in url or "detmold" in url:
                return MockHttpResponse(detail_html)
            raise AssertionError(url)

        with patch.object(scraper.requests, "Session") as session_cls:
            session = session_cls.return_value
            session.get.side_effect = fake_get
            with patch("send_mail.send_excel") as mock_mail:
                scraper.run_scraper()
                mock_mail.assert_called_once()

        wb = load_workbook(tmp_path / "out.xlsx")
        assert wb["Markety"].max_row >= 2
        for row in range(2, wb["Markety"].max_row + 1):
            data_otwarcia = wb["Markety"][f"D{row}"].value
            assert scraper.is_opening_date_in_range(str(data_otwarcia))


class TestOpeningDateFilterIntegration:
    def test_filter_constants_match_requirements(self):
        assert scraper.OPENING_FILTER_START == date(2026, 7, 1)
        assert scraper.OPENING_FILTER_END == date(2028, 12, 31)

    def test_collect_resolves_address_and_filters_dates(
        self, load_fixture, silent_logger, monkeypatch
    ):
        monkeypatch.setattr(scraper, "REQUEST_DELAY_SEC", 0)
        monkeypatch.setattr(scraper, "MAX_PAGES_PER_CATEGORY", 1)

        mixed_html = (
            load_fixture("list_page_linked.html")
            + load_fixture("list_page_out_of_range.html")
        )
        route_map = [
            ("/berlin/zukunft-markt", load_fixture("detail_zukunft_markt.html")),
            ("/moenchengladbach/rewe-esch", load_fixture("detail_rewe_esch.html")),
            ("/frankfurt-am-main/rewe-test", load_fixture("detail_rewe_esch.html")),
            ("/branche/supermaerkte", mixed_html),
        ]
        session = build_mock_session(route_map)

        records = scraper.collect_category_records(
            session=session,
            base_url=f"{scraper.BASE_URL}/branche/supermaerkte",
            cache={},
            logger=silent_logger,
        )

        names = {r.nazwa_firmy for r in records}
        assert "Zukunft Markt" not in names
        assert len(records) == 2
