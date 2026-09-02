"""Testy regresyjne – porównanie wyników parsowania z plikami wzorcowymi (golden files)."""

from __future__ import annotations

from dataclasses import asdict
from datetime import date

import pytest
from bs4 import BeautifulSoup

import neueroeffnung_scraper as scraper

pytestmark = pytest.mark.regression


def _listing_to_dict(item: scraper.ListingItem) -> dict:
    return asdict(item)


class TestRegressionListParsing:
    def test_list_page_linked_matches_golden(self, load_fixture, load_json_fixture):
        expected = load_json_fixture("expected_regression.json")["list_page_linked"]
        html = load_fixture("list_page_linked.html")
        items = scraper.parse_list_page(html, "https://www.neueroeffnung.info/branche/supermaerkte")
        assert [_listing_to_dict(item) for item in items] == expected

    def test_list_page_plain_cards_matches_golden(self, load_fixture, load_json_fixture):
        expected = load_json_fixture("expected_regression.json")["list_page_plain_cards"]
        html = load_fixture("list_page_plain_cards.html")
        items = scraper.parse_list_page(
            html, "https://www.neueroeffnung.info/branche/supermaerkte?page=2"
        )
        assert [_listing_to_dict(item) for item in items] == expected


class TestRegressionDetailParsing:
    @pytest.mark.parametrize(
        "fixture_key,html_file",
        [
            ("detail_rewe_esch", "detail_rewe_esch.html"),
            ("detail_with_closing", "detail_with_closing.html"),
            ("detail_logistik", "detail_logistik.html"),
        ],
    )
    def test_detail_extraction_matches_golden(
        self, fixture_key, html_file, load_fixture, load_json_fixture
    ):
        expected = load_json_fixture("expected_regression.json")[fixture_key]
        soup = BeautifulSoup(load_fixture(html_file), "html.parser")
        actual = {
            "adres": scraper.extract_address(soup),
            "data_otwarcia": scraper.extract_opening_date(soup),
            "data_zamkniecia": scraper.extract_closing_date(
                scraper.extract_information(soup)
            ),
            "informacja": scraper.extract_information(soup),
        }
        assert actual == expected


class TestRegressionConfig:
    def test_config_constants_match_golden(self, load_json_fixture):
        expected = load_json_fixture("expected_regression.json")["config"]
        assert scraper.MAX_ADDRESS_LENGTH == expected["max_address_length"]
        assert scraper.MAX_INFO_LENGTH == expected["max_info_length"]
        assert scraper.OPENING_FILTER_START.isoformat() == expected["opening_filter_start"]
        assert scraper.OPENING_FILTER_END.isoformat() == expected["opening_filter_end"]
        assert list(scraper.EXCEL_COLUMNS) == expected["excel_columns"]
        assert list(scraper.DATA_SHEET_NAMES) == expected["data_sheet_names"]

    def test_opening_date_samples_match_golden(self, load_json_fixture):
        samples = load_json_fixture("expected_regression.json")["opening_date_samples"]
        for text in samples["in_range"]:
            assert scraper.is_opening_date_in_range(text) is True
        for text in samples["out_of_range"]:
            assert scraper.is_opening_date_in_range(text) is False


class TestRegressionRecordShape:
    def test_record_fields_stable(self):
        record = scraper.Record(
            nazwa_firmy="Test",
            adres="Adres 1",
            data_zamkniecia="01.01.2026",
            data_otwarcia="01.06.2026",
        )
        assert tuple(asdict(record).keys()) == (
            "nazwa_firmy",
            "adres",
            "data_zamkniecia",
            "data_otwarcia",
            "informacja",
            "typ_wpisu",
            "kategoria",
            "detail_url",
            "listing_adres_lista",
            "entry_type_raw",
            "status_walidacji",
            "brakujace_pola",
            "proby_ponowienia",
            "godziny_pracy",
            "maps_zweryfikowany",
        )

    def test_excel_column_order_unchanged(self, silent_logger, tmp_path):
        output_path = tmp_path / "regression.xlsx"
        scraper.write_excel(
            {"Markets": [scraper.Record("A", "B", "C", "D")]},
            [],
            output_path,
            silent_logger,
        )
        from openpyxl import load_workbook

        wb = load_workbook(output_path)
        ws = wb["Markets"]
        assert [ws.cell(row=1, column=i).value for i in range(1, 9)] == list(
            scraper.EXCEL_COLUMNS
        )

    def test_address_validation_respects_650_limit(self, silent_logger):
        long_adres = "X" * 700
        result = scraper.validate_address(long_adres, silent_logger)
        assert len(result) == 650
