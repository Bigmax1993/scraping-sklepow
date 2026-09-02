"""Testy kolumn kontaktowych w eksporcie Excel."""

from __future__ import annotations

from openpyxl import load_workbook

import neueroeffnung_scraper as scraper


def test_record_to_excel_row_includes_contact_fields():
    record = scraper.Record(
        nazwa_firmy="REWE Esch",
        adres="Beckrather Straße 39, 41189 Mönchengladbach",
        data_zamkniecia="",
        data_otwarcia="03.09.2026",
        telefon="+49 2161 12345",
        email="info@rewe-esch.de",
        osoba_kontaktowa="Max Mustermann",
    )

    row = scraper.record_to_excel_row(record)

    assert row[-3:] == ["+49 2161 12345", "info@rewe-esch.de", "Max Mustermann"]


def test_record_to_excel_row_keeps_empty_contacts_when_missing():
    record = scraper.Record(
        nazwa_firmy="Lidl",
        adres="Hauptstraße 1, 32756 Detmold",
        data_zamkniecia="",
        data_otwarcia="09.09.2026",
    )

    row = scraper.record_to_excel_row(record)

    assert row[-3:] == ["", "", ""]


def test_write_excel_exports_contact_columns(silent_logger, tmp_path):
    output_path = tmp_path / "contacts.xlsx"
    sheets = {
        "Markets": [
            scraper.Record(
                "REWE Esch",
                "Beckrather Straße 39, 41189 Mönchengladbach",
                "",
                "03.09.2026",
                telefon="+49 2161 12345",
                email="info@rewe-esch.de",
                osoba_kontaktowa="Max Mustermann",
            ),
            scraper.Record(
                "Lidl",
                "Hauptstraße 1, 32756 Detmold",
                "",
                "09.09.2026",
            ),
        ],
        "Restaurants": [],
        "Drugstores": [],
        "Shopping centers": [],
    }

    scraper.write_excel(sheets, [], output_path, silent_logger)

    wb = load_workbook(output_path)
    ws = wb["Markets"]
    assert ws["I1"].value == "Phone"
    assert ws["J1"].value == "Email"
    assert ws["K1"].value == "Contact person"
    assert ws["I2"].value == "+49 2161 12345"
    assert ws["J2"].value == "info@rewe-esch.de"
    assert ws["K2"].value == "Max Mustermann"
    assert ws["I3"].value in ("", None)
    assert ws["J3"].value in ("", None)
    assert ws["K3"].value in ("", None)
